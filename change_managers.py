#change_managers.py
''' script to change managers for a group of users
  requires an excel file - File format:
  ID,ANY,ANYManager ID
  00010000,any,any,0100010
  00010000,any,any,0100010
  00010000,any,any,0100010
  ...
  COMMAND LINE ARGUMENTS
  -i <filenmae>: Input XLSX file of User Id's and their NEW manager ID's
  -s <filename>: Saved LDIF of the users before the changes
  -n <filename>: Saved LIDF of the users AFTER the changes
'''
import sys
from ldap_client import *
from ldap3 import MODIFY_REPLACE, SUBTREE
import time
import openpyxl 

def print_help():
    print('''Invalid Operation:
python change_managers.py -i<input xlsx fiename> -b <ldif before changes> 
    -a <ldif after changes> 
    ''')
    print('LDIF files stored in <path_to_local_dir>/data directory')
    print('Example: python3 -i change_managers.py -b p1.xslx -a p1_before.ldif p2_after.ldif')

def get_command_line_args():
    #print(f'Number of arguments:  {len(sys.argv)} arguments.') 
    #print(f'Argument List:  {str(sys.argv)}')
    # required 3 arguments, option 4th argument
    if len(sys.argv) < 4:
        print_help()
        exit(1)
    input_xlsx, before_ldif, after_ldif = None, None, None
    dir_server = 'enterprise'
    i = 1
    while i < len(sys.argv):
        cmd = sys.argv[i]
        val = sys.argv[i+1]
        if cmd == '-i':
            input_xlsx = val
        if cmd == '-b':
            before_ldif = val
        if cmd == '-a':
            after_ldif = val 
        i += 2
    print(f'Input File: {input_xlsx}, before ldif: {before_ldif}')
    print(f'after ldif: {after_ldif}, dir server: {dir_server}')
    if None in (input_xlsx, before_ldif, after_ldif):
        print_help()
        exit(1)
    return input_xlsx, before_ldif, after_ldif, dir_server

def pad_zeros(n):
    # make sure format is 8 digits, padd zeros as necessary
    s = str(n)
    s2 = s.rjust(8, '0')
    return s2

def get_changes_from_xslx_file(input_xlsx):
    ids = []
    manager_ids = []
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    # top row should have 4 values, namely: 
    #   'BmsId', 'Name', 'New Manager Name', 'New Manager BMSID'
    top_row = [ws.cell(row=1,column=i).value for i in range(1,5)]
    # print(top_row)
    expected_top_row = ['BmsId', 'Name', 'New Manager Name', 'New Manager BMSID']
    if top_row != expected_top_row:
        print('Invalid file contents - STOPPING!')
    # read row by row until an empty cell is found
    row = 2
    while True:
        next_row = [ws.cell(row=row,column=i).value for i in range(1,5)]
        if (next_row[0] == None) or (next_row[0] == '') or (next_row[0] == ' '):
            break
        #print(next_row)
        ids.append(pad_zeros(next_row[0]))
        manager_ids.append(pad_zeros(next_row[3]))
        row += 1
    return ids,manager_ids
    # with open(input_csv,'r') as f1:
    #     lines = f1.readlines()
    # ids = []
    # manager_ids = []
    # for line in lines:
    #     line = line.strip()
    #     line_split = line.split(',')
    #     if 'id' in line_split[0].lower():
    #         pass
    #     else:
    #         ids.append(line_split[0])
    #         manager_ids.append(line_split[1])
    # return ids,manager_ids

def save_to_ldif(filename, bms_ids, dir_creds, search_base):
    # break into groups of 10, so we can save to a single ldif file
    start = 0
    end = len(bms_ids)
    step = 4
    append_file = False # first time write to file, next time appened it
    for i in range(start, end, step):
        last = i + step
        if last > end:
            last = end
        #print(f'save {i},{i+step}')
        s = '(|'
        for id in range(i,last):
            s += '(bmsid=' + bms_ids[id] + ')'
        s += ')'
        #print('->filter:',s)
        search_filter_all_users = s
        search_scope = SUBTREE
        _,results,ldif = ldap_search(dir_creds, search_base, 
           search_filter_all_users, search_scope, '*',get_ldif=True)
        ldif_to_file(filename,ldif,append_file)
        append_file = True

# now make the change
def make_changes(bms_ids, manager_ids, dir_creds):
    search_filter = []
    modify_user_dn_enterprise = []
    changes_manager = []
    for i in range(len(bms_ids)):
        search_filter.append('(bmsid=' + bms_ids[i] + ')')
        modify_user_dn_enterprise.append('bmsid=' + bms_ids[i] + ',ou=People,o=bms.com')
        change_mgr = 'bmsid=' + manager_ids[i] + ',ou=People,o=bms.com'
        changes_manager.append({
            'bmsmanagersid': [(MODIFY_REPLACE, [manager_ids[i]])],
            'manager': [(MODIFY_REPLACE, [change_mgr])],
            })

    for i in range(len(bms_ids)):
        print(f'Change manager for bmsid={bms_ids[i]}')
        search_scope = SUBTREE
        _ = modify_ldap_user(dir_creds, modify_user_dn_enterprise[i], changes_manager[i], None)
        # check it was changed
        _,results = ldap_search(dir_creds, search_base_enterprise, search_filter[i], search_scope, 'bmsmanagersid')
        print(results)
        _,r = get_attr_value_if_exists(results, 'bmsmanagersid')
        print(f'User {bms_ids[i]} new bmsmanagerid {r}')
        _,results = ldap_search(dir_creds, search_base_enterprise, search_filter[i], search_scope, 'manager')
        print(results)
        _,r = get_attr_value_if_exists(results, 'manager')
        print(f'User {bms_ids[i]} new manager {r}')


#--- COMMAND LINE PROCESSING ---#
input_xlsx, before_ldif, after_ldif, dir_server = get_command_line_args()
ids, manager_ids = get_changes_from_xslx_file(input_xlsx)
#print(ids,manager_ids)
#exit(0)
dir_creds = get_creds_from_server(dir_server)
if dir_server == 'enterprise':
    search_base_enterprise  = 'o=bms.com'
else:
    print(f'SERVER {dir_server} not supported yet')
    exit(1)

#----- MAIN CODE -----#
# save before changes to ldif file
save_to_ldif(before_ldif, ids, dir_creds, search_base_enterprise)
cleanup_lidf_file(before_ldif)
print(f'Before changes saved to data/{before_ldif}')
# make the changes
make_changes(ids, manager_ids, dir_creds)
print('Changes COMPLETE')
# save the chagnes
save_to_ldif(after_ldif, ids, dir_creds, search_base_enterprise)
cleanup_lidf_file(after_ldif)
print(f'After changes saved to data/{after_ldif}')
