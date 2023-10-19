#celgene_report.py
''' script to change report celgene users in enterprise directory
  requires an excel file - File format:
  hCel UserID
  USERA
  USERB
  ...
  COMMAND LINE ARGUMENTS
  -i <filenmae>: Input XLSX file of Celgene users
'''
import sys
from ldap_client import *
from ldap3 import SUBTREE
import time
import openpyxl 

missing_from_celgene = []
ext2_missing_from_celgene = []
bmsid_missing_from_ent = []

def print_help():
    print('''Invalid Operation:
python3 celgene_report.py -i<input xlsx fiename> -o<output xlsx filename>
    ''')
    print('Example: python3 celgene_report.py -i p1.xlsx -o p2.xlsx')

def get_command_line_args():
    #print(f'Number of arguments:  {len(sys.argv)} arguments.') 
    #print(f'Argument List:  {str(sys.argv)}')
    # required 2 arguments
    if len(sys.argv) < 3:
        print_help()
        exit(1)
    input_xlsx,output_xlsx = None,None
    dir_server = 'enterprise'
    i = 1
    while i < len(sys.argv):
        cmd = sys.argv[i]
        val = sys.argv[i+1]
        if cmd == '-i':
            input_xlsx = val
        if cmd == '-o':
            output_xlsx = val
        i += 2
    if (input_xlsx == None) or (output_xlsx==None):
        print_help()
        exit(1)
    return input_xlsx, output_xlsx, dir_server

def get_users_from_xslx_file(input_xlsx):
    ids = []
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    # top row should have 4 values, namely: 
    top_row = [ws.cell(row=1,column=i).value for i in range(1,2)]
    # print(top_row)
    expected_top_row = ['hCel UserID']
    if top_row != expected_top_row:
        print('Invalid file contents - STOPPING!')
    # read row by row until an empty cell is found
    row = 2
    while True:
        next_row = [ws.cell(row=row,column=i).value for i in range(1,2)]
        if (next_row[0] == None) or (next_row[0] == '') or (next_row[0] == ' '):
            break
        #print(next_row)
        ids.append(next_row[0])
        row += 1
    return ids

def get_row_from_enterprise(employeeid,bmsid=None):
    global bmsid_missing_from_ent
    if bmsid:
        search_filter = f'(bmsid={bmsid})'
        #print(f'->FIND {bmsid}')
    else:
        search_filter = f'(employeeid={employeeid})'
        # print(f'->FIND {employeeid}')
    attrs = ['bmsid','uid','mail','cn','bmsentaccountstatus']
    _,results = ldap_search(dir_creds, search_base_enterprise, search_filter, search_scope, attrs)
    if (not results) and bmsid:
        bmsid_missing_from_ent.append(employeeid)
        return None
    # print the output row
    if results and (len(results) > 0):
        d = result_to_dict(results)
        if 'mail' not in d:
            d['mail'] = ''
        # format employeeid,bmsid,mail,cn,bmsentaccoutstatus
        output_row = [employeeid,d["bmsid"],d["uid"],d["mail"],d["cn"],d["bmsentaccountstatus"]]
        return output_row
    return None # not found

def get_extensionattribute2_from_celgene(samaccountname): 
    global missing_from_celgene, ext2_missing_from_celgene 
    search_filter = f'(samaccountname={samaccountname})'
    attrs = ['samaccountname','extensionAttribute2','extensionattribute6']
    _,results = ldap_search(celgene_creds, search_base_celgene, search_filter, search_scope, attrs)
    if not results:
        missing_from_celgene.append(samaccountname)
        return None
    r,ext2 = get_attr_value_if_exists(results,'extensionAttribute2')    
    if (not r) or (not ext2):
        print(f'missing EXT2!! {samaccountname}')
        ext2_missing_from_celgene.append(samaccountname)
        return None
    return ext2

def process_row(id,sheet):
    # try enterprise first
    print('.',end='',flush=True)
    row = get_row_from_enterprise(id)
    if row:
        #print(row)
        sheet.append(row)
        return
    # try celgene to get bmsid
    #print(f'did not find {id} in enterprise')
    ext2 = get_extensionattribute2_from_celgene(id)
    if ext2:
        row = get_row_from_enterprise(id,ext2)
        if row:
            #print(row)
            sheet.append(row)
            return
    # did not find, just write the first column
    row = [id]
    sheet.append(row)
    
# now do the report
def make_report(ids,dir_creds,output_xlsx,celgene_creds):
    book = openpyxl.Workbook()
    sheet = book.active
    top_row = ['celgene user id','bmsid','bms uid','mail','cn','bmsentaccoutstatus']
    sheet.append(top_row)   
    for id in ids:
        process_row(id,sheet)
    book.save(output_xlsx)

#--- COMMAND LINE PROCESSING ---#
input_xlsx, output_xlsx, dir_server = get_command_line_args()
ids = get_users_from_xslx_file(input_xlsx)
#exit(0)
dir_creds = get_creds_from_server(dir_server)
if dir_server == 'enterprise':
    search_base_enterprise  = 'o=bms.com'
else:
    print(f'SERVER {dir_server} not supported yet')
    exit(1)
search_scope = SUBTREE
# need celgene for some ingo
#celgene_creds = get_creds_from_server('uslvlspadsadc01')
# format uslvlspadsadc01.celgene.com||CN=svc_pcim,OU=Service Accounts,OU=Restricted,DC=celgene,DC=com|389
celgene_creds = [
'ussumspadsadc01.celgene.com',
'j01n3ng1n3',
'CN=SVC_CEL_JE,OU=Service Accounts,OU=Users,OU=Infrastructure OU,DC=celgene,DC=com',
636
]

#print(celgene_creds)
#exit(0)
search_base_celgene = 'DC=celgene,DC=com'

#----- MAIN CODE -----#
# make the report
make_report(ids,dir_creds,output_xlsx,celgene_creds)
print(f'\nmissing_from_celgene {missing_from_celgene}')
print(f'ext2_missing_from_celegene {ext2_missing_from_celgene}')
print(f'bmsid_missing_from_ent {bmsid_missing_from_ent}')
print('DONE')
